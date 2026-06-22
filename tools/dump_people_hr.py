"""
Dump the full People HR employee dataset to xlsx so you can see exactly
what the API returns.

Hits POST /Employee with Action=GetAllEmployeeDetail twice - once with
IncludeLeavers=false (what the app uses), once with IncludeLeavers=true
- so you can see if leavers are being filtered out at source.

Outputs:
  - Console summary (count, by Status, by Branch, by JobRole)
  - people_hr_dump.xlsx with one sheet per fetch
"""
import sys, pathlib
sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent.parent))

from collections import Counter

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from config import (
    PEOPLE_HR_BASE_URL,
    PEOPLE_HR_API_KEY,
    PEOPLE_HR_TIMEOUT,
    PEOPLE_HR_EMPLOYEE_RESOURCE,
    PEOPLE_HR_ACTION_EMPLOYEES,
)

OUT_DEFAULT = "people_hr_dump.xlsx"
OUT = sys.argv[1] if len(sys.argv) >= 2 else OUT_DEFAULT

# Fields to unwrap to flat columns. People HR wraps each field as
# {DisplayValue: "..."} - this list controls what makes it into the dump.
TOP_FIELDS = [
    "EmployeeId", "Title", "FirstName", "LastName", "OtherName", "KnownAs",
    "EmailId", "StartDate", "DateOfBirth", "JobRole", "Company",
    "CompanyEffectiveDate", "Location", "LocationEffectiveDate",
    "Department", "DepartmentEffectiveDate", "JobRoleChangeDate",
    "ReportsTo", "ReportsToEmployeeId", "ReportsToEmailAddress",
    "NISNumber", "Nationality", "EmploymentType", "EmployeeStatus",
    "HolidayAllowanceDays", "HolidayAllowanceMins", "NoticePeriod",
    "ProbationEndDate", "Gender",
]

# Some leaver-specific top-level fields (not wrapped)
LEAVER_FIELDS = ["LeavingDate", "ReasonForLeaving"]


def _dv(field):
    """Unwrap People HR's {DisplayValue: ...} envelope (returns '' if absent)."""
    if isinstance(field, dict):
        return str(field.get("DisplayValue", "") or "").strip()
    if field is None:
        return ""
    return str(field).strip()


def fetch(include_leavers):
    url = PEOPLE_HR_BASE_URL.rstrip("/") + PEOPLE_HR_EMPLOYEE_RESOURCE
    payload = {
        "APIKey": PEOPLE_HR_API_KEY,
        "Action": PEOPLE_HR_ACTION_EMPLOYEES,
        "IncludeLeavers": "true" if include_leavers else "false",
    }
    print(f"POST {url}    IncludeLeavers={payload['IncludeLeavers']}")
    resp = requests.post(url, json=payload, timeout=PEOPLE_HR_TIMEOUT)
    print(f"  HTTP {resp.status_code}    bytes={len(resp.content)}")
    resp.raise_for_status()
    data = resp.json()
    is_error = str(data.get("IsError", "")).lower() == "true"
    msg = data.get("Message", "")
    print(f"  IsError={is_error}    Message={msg!r}")
    if is_error:
        raise RuntimeError(f"People HR error: {msg}")
    return data.get("Result") or []


def flatten(rec):
    out = {}
    for f in TOP_FIELDS:
        out[f] = _dv(rec.get(f))
    for f in LEAVER_FIELDS:
        v = rec.get(f)
        out[f] = "" if v is None else str(v).strip()
    return out


def write_sheet(wb, title, records):
    ws = wb.create_sheet(title=title[:31])
    cols = TOP_FIELDS + LEAVER_FIELDS
    # Header
    fill = PatternFill("solid", fgColor="305496")
    font = Font(bold=True, color="FFFFFF")
    for ci, c in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(ci)].width = 18
    # Body
    for ri, rec in enumerate(records, start=2):
        flat = flatten(rec)
        for ci, c in enumerate(cols, 1):
            ws.cell(row=ri, column=ci, value=flat[c])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def summarize(label, records):
    print(f"\n=== {label}: {len(records)} records ===")
    if not records:
        return
    flat = [flatten(r) for r in records]

    def topn(field, n=15):
        c = Counter((f.get(field) or "(blank)") for f in flat)
        print(f"  Top {field} values:")
        for v, cnt in c.most_common(n):
            print(f"    {cnt:>4}  {v!r}")

    topn("EmployeeStatus", 5)
    topn("Company", 8)
    topn("Location", 12)
    topn("Department", 8)
    topn("JobRole", 12)
    # Count with/without LeavingDate
    with_leave = sum(1 for f in flat if f["LeavingDate"])
    print(f"  Records with LeavingDate set: {with_leave}")


def main():
    if not PEOPLE_HR_API_KEY:
        print("ERROR: PEOPLE_HR_API_KEY is empty in config.py")
        sys.exit(1)

    print("--- Fetch 1: IncludeLeavers=false (what the app uses) ---")
    no_leavers = fetch(include_leavers=False)
    summarize("IncludeLeavers=false", no_leavers)

    print("\n--- Fetch 2: IncludeLeavers=true (everything) ---")
    with_leavers = fetch(include_leavers=True)
    summarize("IncludeLeavers=true", with_leavers)

    # Show the delta
    print(f"\n--- Delta ---")
    print(f"  IncludeLeavers=false: {len(no_leavers):>4} records")
    print(f"  IncludeLeavers=true:  {len(with_leavers):>4} records")
    diff = len(with_leavers) - len(no_leavers)
    print(f"  Difference (leavers): {diff:>4}")

    # Write xlsx
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    write_sheet(wb, "no_leavers", no_leavers)
    write_sheet(wb, "all_with_leavers", with_leavers)

    out_path = pathlib.Path(__file__).resolve().parent.parent / OUT
    wb.save(out_path)
    print(f"\nWrote {out_path}")


if __name__ == "__main__":
    main()
