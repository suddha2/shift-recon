"""
Filter the CSV to rows for one employee (matched against both Actual
and Planned Employee Name) and write to .xlsx. Date columns are written
as real Excel datetimes with a DD/MM/YYYY HH:MM format, so Excel doesn't
display them as serial numbers.

Adds derived columns at the front: _Row (CSV row number), _ISO_Year,
_ISO_Week, _Hours so the user can sort/filter easily.

Usage: python export_emp.py
"""
import re
import sys, pathlib
sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent.parent))

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from analyzer import parse_datetime, get_week_number, calculate_hours

CSV = "Supported Living Recon V5 (3).csv"
EMP_QUERY = "Akabogu, Genevieve"
OUT = "Akabogu_Genevieve_all_rows.xlsx"

DATE_FORMAT = "DD/MM/YYYY HH:MM"


def read_csv(path):
    for enc in (None, "cp1252", "latin-1"):
        try:
            kw = {"low_memory": False}
            if enc is not None:
                kw["encoding"] = enc
            return pd.read_csv(path, **kw)
        except UnicodeDecodeError:
            continue
    raise RuntimeError(f"Could not decode {path}")


def looks_like_date_col(name):
    """True for columns whose values are typically date/datetime strings."""
    lower = name.lower()
    # Match 'date', 'time', or both, but exclude e.g. 'PunctualityStartTimeMinutes'
    if "minutes" in lower:
        return False
    return ("date" in lower) or ("time" in lower)


def main():
    df = read_csv(CSV)
    df.columns = df.columns.str.strip().str.replace(r"\s+", " ", regex=True)
    df.columns = df.columns.str.replace("Desciption", "Description", regex=False)
    df.columns = df.columns.str.replace("and Time", "And Time", regex=False)
    df["_Row"] = range(2, len(df) + 2)

    actual_name = df["Actual Employee Name"].astype(str)
    planned_name = df["Planned Employee Name"].astype(str)
    mask = (
        actual_name.str.contains(EMP_QUERY, case=False, na=False)
        | planned_name.str.contains(EMP_QUERY, case=False, na=False)
    )
    sub = df[mask].copy()

    # Derived helpers — ISO week, hours
    sub["_ParsedStart"] = sub["Actual Start Date And Time"].apply(parse_datetime)
    sub["_ParsedEnd"]   = sub["Actual End Date And Time"].apply(parse_datetime)
    sub["_Hours"]       = sub.apply(lambda r: round(calculate_hours(r["_ParsedStart"], r["_ParsedEnd"]), 2), axis=1)
    sub["_ISO_Week"]    = sub["_ParsedStart"].apply(lambda x: get_week_number(x)[0])
    sub["_ISO_Year"]    = sub["_ParsedStart"].apply(lambda x: get_week_number(x)[1])
    sub = sub.sort_values("_ParsedStart")
    sub = sub.drop(columns=["_ParsedStart", "_ParsedEnd"])

    # Move derived columns to the front
    derived = ["_Row", "_ISO_Year", "_ISO_Week", "_Hours"]
    other = [c for c in sub.columns if c not in derived]
    sub = sub[derived + other]

    # Identify date columns and pre-parse them to real datetimes so Excel
    # treats them as dates (not serial numbers).
    date_cols = [c for c in other if looks_like_date_col(c)]
    for col in date_cols:
        sub[col] = sub[col].apply(parse_datetime)

    # Write via openpyxl directly so we can apply the date number format
    wb = Workbook()
    ws = wb.active
    ws.title = re.sub(r"[^\w]+", "_", EMP_QUERY)[:31]

    # Header
    for ci, col in enumerate(sub.columns, start=1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")

    # Body
    for ri, (_, row) in enumerate(sub.iterrows(), start=2):
        for ci, col in enumerate(sub.columns, start=1):
            v = row[col]
            if pd.isna(v):
                ws.cell(row=ri, column=ci, value=None)
                continue
            cell = ws.cell(row=ri, column=ci, value=v)
            if col in date_cols:
                cell.number_format = DATE_FORMAT

    # Freeze header row + tidy column widths (rough)
    ws.freeze_panes = "A2"
    for ci, col in enumerate(sub.columns, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = max(10, min(28, len(str(col)) + 2))

    wb.save(OUT)
    print(f"Wrote {len(sub)} rows to {OUT}")
    print(f"Sheet: {ws.title}")
    print(f"Date columns formatted as {DATE_FORMAT}: {len(date_cols)}")


if __name__ == "__main__":
    main()
