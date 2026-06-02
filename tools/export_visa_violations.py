"""
Run the visa-hour check and export the full violation list to xlsx so
the operator has everything in one place (no truncation, sortable,
filterable in Excel).
"""
import sys, pathlib
sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent.parent))

import io

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from analyzer import check_visa_hour_violations, canonical_name
from config import (
    APP_EMP_URL, APP_EMP_AUTH, APP_EMP_TIMEOUT,
    APP_EMP_NAME_COL, APP_EMP_TYPE_COL,
)

CSV = "/mnt/c/Users/SadharsunRamalingma/Downloads/Supported Living Recon V5 (2)-redone.csv"
OUT = "visa_violations_full.xlsx"


def read_csv(path):
    for enc in (None, "cp1252", "latin-1"):
        try:
            kw = {"low_memory": False}
            if enc is not None:
                kw["encoding"] = enc
            return pd.read_csv(path, **kw)
        except UnicodeDecodeError:
            continue
    raise RuntimeError(f"Could not decode {path}")


def fetch_visa_lookup():
    resp = requests.get(APP_EMP_URL, headers={"Authorization": APP_EMP_AUTH},
                        timeout=APP_EMP_TIMEOUT)
    resp.raise_for_status()
    feed = pd.read_csv(io.StringIO(resp.text), dtype=str)
    feed.columns = feed.columns.str.strip()
    lookup = {}
    for _, row in feed.iterrows():
        name = str(row.get(APP_EMP_NAME_COL, "") or "").strip().replace("-", " ")
        emp_type = str(row.get(APP_EMP_TYPE_COL, "") or "").strip()
        visa_status = emp_type.split(" - ", 1)[1].strip() if " - " in emp_type else ""
        if name and name.lower() != "nan":
            lookup[canonical_name(name)] = visa_status
    return lookup


def main():
    df = read_csv(CSV)
    df.columns = df.columns.str.strip().str.replace(r"\s+", " ", regex=True)
    df.columns = df.columns.str.replace("Desciption", "Description", regex=False)
    df.columns = df.columns.str.replace("and Time", "And Time", regex=False)
    df["_row_num"] = range(2, len(df) + 2)

    visa_lookup = fetch_visa_lookup()
    violations = check_visa_hour_violations(df, visa_lookup)

    # Flatten to rows with derived columns
    rows = []
    for v in violations:
        details = str(v.get("details", "")).lower()
        if v["issue_type"] == "Missing Visa Info":
            kind = "Missing Visa"
            gap = None
        elif "exceeds maximum" in details:
            kind = "OVER"
            try:
                gap = round(float(v["actual_hours"]) - float(v["limit_hours"]), 1)
            except (TypeError, ValueError):
                gap = None
        elif "below minimum" in details:
            kind = "UNDER"
            try:
                gap = round(float(v["limit_hours"]) - float(v["actual_hours"]), 1)
            except (TypeError, ValueError):
                gap = None
        else:
            kind = "OTHER"
            gap = None

        rows.append({
            "Employee": v.get("employee_name", ""),
            "Week": v.get("week", ""),
            "Week Range": v.get("date", ""),
            "Type": kind,
            "Visa Status": v.get("shift_type", ""),
            "Actual Hours": v.get("actual_hours"),
            "Limit Hours": v.get("limit_hours"),
            "Gap (h)": gap,
            "Source CSV Row(s)": v.get("row_numbers", ""),
            "Details": v.get("details", ""),
        })

    # Sort: Over first by largest gap, then Under by largest gap, then Missing
    def sort_key(r):
        kind_order = {"OVER": 0, "UNDER": 1, "OTHER": 2, "Missing Visa": 3}
        return (
            kind_order.get(r["Type"], 9),
            -(r["Gap (h)"] or 0),
            r["Employee"],
            r["Week"],
        )
    rows.sort(key=sort_key)

    # Build xlsx
    wb = Workbook()
    ws = wb.active
    ws.title = "Visa Hour Violations"

    headers = ["Employee", "Week", "Week Range", "Type", "Visa Status",
               "Actual Hours", "Limit Hours", "Gap (h)",
               "Source CSV Row(s)", "Details"]
    widths  = [30, 10, 24, 12, 24, 13, 12, 9, 30, 90]

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header row
    fill = PatternFill("solid", fgColor="305496")
    font = Font(bold=True, color="FFFFFF")
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
        ws.column_dimensions[get_column_letter(ci)].width = widths[ci - 1]

    # Type colour bands
    over_fill   = PatternFill("solid", fgColor="FCE4D6")
    under_fill  = PatternFill("solid", fgColor="FFF2CC")
    miss_fill   = PatternFill("solid", fgColor="EDEDED")

    for ri, r in enumerate(rows, start=2):
        row_fill = (over_fill if r["Type"] == "OVER"
                    else under_fill if r["Type"] == "UNDER"
                    else miss_fill if r["Type"] == "Missing Visa"
                    else None)
        for ci, h in enumerate(headers, 1):
            v = r[h]
            c = ws.cell(row=ri, column=ci, value=v)
            c.border = border
            if row_fill:
                c.fill = row_fill
            if h in ("Actual Hours", "Limit Hours", "Gap (h)"):
                c.alignment = Alignment(horizontal="right")
                c.number_format = "0.0"
            else:
                c.alignment = Alignment(vertical="top", wrap_text=False)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Summary sheet
    sh = wb.create_sheet("Summary")
    counts = {"OVER": 0, "UNDER": 0, "OTHER": 0, "Missing Visa": 0}
    visa_counts = {}
    for r in rows:
        counts[r["Type"]] = counts.get(r["Type"], 0) + 1
        v = r["Visa Status"]
        visa_counts[v] = visa_counts.get(v, 0) + 1

    sh["A1"] = "Visa Hour Violations - Summary"
    sh["A1"].font = Font(bold=True, size=14)
    sh["A3"] = "By Type"
    sh["A3"].font = Font(bold=True)
    sh["A4"] = "Type";       sh["B4"] = "Count"
    for c in (sh["A4"], sh["B4"]): c.font = Font(bold=True)
    for i, (k, v) in enumerate(counts.items(), start=5):
        sh.cell(row=i, column=1, value=k)
        sh.cell(row=i, column=2, value=v)
    sh.cell(row=5 + len(counts), column=1, value="TOTAL").font = Font(bold=True)
    sh.cell(row=5 + len(counts), column=2, value=sum(counts.values())).font = Font(bold=True)

    sh["D3"] = "By Visa Status"
    sh["D3"].font = Font(bold=True)
    sh["D4"] = "Visa";       sh["E4"] = "Count"
    for c in (sh["D4"], sh["E4"]): c.font = Font(bold=True)
    for i, (k, v) in enumerate(sorted(visa_counts.items(), key=lambda x: -x[1]), start=5):
        sh.cell(row=i, column=4, value=k or "(none)")
        sh.cell(row=i, column=5, value=v)

    sh.column_dimensions["A"].width = 16
    sh.column_dimensions["B"].width = 10
    sh.column_dimensions["D"].width = 28
    sh.column_dimensions["E"].width = 10

    out_path = pathlib.Path(__file__).resolve().parent.parent / OUT
    wb.save(out_path)
    print(f"Wrote {out_path}")
    print(f"Total violations: {len(rows)}")
    for k, v in counts.items():
        if v:
            print(f"  {k}: {v}")


if __name__ == "__main__":
    main()
