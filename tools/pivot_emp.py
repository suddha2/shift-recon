"""
Per-employee weekly pivot in xlsx, matching the sample layout:

  Row 1: "Emp: <name>" (title)
  Row 2: "" | "" | Week 1 (merged 7 cols) | Week 2 | Week 3 | Week 4
  Row 3: Service | Shift Type | <7 daily dates> | <7> | <7> | <7>
  Row 4+: one row per (Service Location, Shift Type) combination,
          cells = sum of Actual Hours on that calendar day

Week boundaries: 4 contiguous 7-day blocks, with Week 1 starting at the
MIN Actual Start Date across the WHOLE CSV (not the employee).
Shifts crossing midnight are bucketed under the START date.

Usage:
  python pivot_emp.py                       # defaults to Akabogu, Genevieve
  python pivot_emp.py "Surname, Firstname"
"""
import re
import sys, pathlib
sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent.parent))

from collections import defaultdict
from datetime import timedelta

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from analyzer import parse_datetime, calculate_hours

CSV = "Supported Living Recon V5 (2).csv"
EMP = sys.argv[1] if len(sys.argv) >= 2 else "Akabogu, Genevieve"
N_WEEKS = 4

# Styling
WEEK_HEADER_FILL = PatternFill("solid", fgColor="305496")
WEEK_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUB_HEADER_FILL  = PatternFill("solid", fgColor="D9E1F2")
SUB_HEADER_FONT  = Font(bold=True)
TITLE_FONT       = Font(bold=True, size=12)
TOTAL_FILL       = PatternFill("solid", fgColor="FFF2CC")
TOTAL_FONT       = Font(bold=True)
THIN             = Side(style="thin", color="CCCCCC")
CELL_BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def read_csv(path):
    for enc in (None, "cp1252", "latin-1"):
        try:
            kw = {"low_memory": False}
            if enc is not None:
                kw["encoding"] = enc
            return pd.read_csv(path, **kw)
        except UnicodeDecodeError:
            continue
    raise RuntimeError(f"Could not decode {path}")


def main():
    df = read_csv(CSV)
    df.columns = df.columns.str.strip().str.replace(r"\s+", " ", regex=True)
    df.columns = df.columns.str.replace("Desciption", "Description", regex=False)
    df.columns = df.columns.str.replace("and Time", "And Time", regex=False)

    # Determine Week 1 start = min Actual Start Date across whole CSV
    all_starts = df["Actual Start Date And Time"].apply(parse_datetime).dropna()
    if all_starts.empty:
        print("No parseable Actual Start dates in the CSV")
        sys.exit(1)
    week1_start = all_starts.min().date()
    week_dates = [week1_start + timedelta(days=i) for i in range(N_WEEKS * 7)]
    period_end = week_dates[-1]
    print(f"Period covered by pivot: {week1_start} (Week 1, Day 1) -> {period_end} (Week {N_WEEKS}, Day 7)")

    # Filter to the employee (actual OR planned name match)
    a = df["Actual Employee Name"].astype(str)
    p = df["Planned Employee Name"].astype(str)
    sub = df[
        a.str.contains(EMP, case=False, na=False)
        | p.str.contains(EMP, case=False, na=False)
    ].copy()
    if sub.empty:
        print(f"No rows matched '{EMP}'")
        sys.exit(1)

    sub["_start"] = sub["Actual Start Date And Time"].apply(parse_datetime)
    sub["_end"]   = sub["Actual End Date And Time"].apply(parse_datetime)
    sub["_hours"] = sub.apply(lambda r: round(calculate_hours(r["_start"], r["_end"]), 2), axis=1)
    sub["_date"]  = sub["_start"].apply(
        lambda x: x.date() if (x is not None and not pd.isna(x)) else None
    )

    # Keep only rows within the 4-week window
    sub = sub[sub["_date"].apply(
        lambda d: d is not None and not pd.isna(d) and week_dates[0] <= d <= week_dates[-1]
    )]
    if sub.empty:
        print(f"'{EMP}' has no activity in the 4-week window {week_dates[0]} - {week_dates[-1]}")
        sys.exit(1)

    # Pivot: (Service Location, Shift Type) x date -> sum(hours)
    pivot = defaultdict(lambda: defaultdict(float))
    for _, r in sub.iterrows():
        location = (str(r.get("Service Location Name", "") or "")).strip() or "(blank)"
        shift_type = (str(r.get("Actual Service Type Description", "") or "")).strip() or "(blank)"
        key = (location, shift_type)
        pivot[key][r["_date"]] += r["_hours"]

    # Sort rows by Service then Shift Type
    sorted_keys = sorted(pivot.keys(), key=lambda k: (k[0].lower(), k[1].lower()))

    # Build workbook
    wb = Workbook()
    ws = wb.active
    ws.title = re.sub(r"[^\w]+", "_", EMP)[:31]

    total_cols = 2 + N_WEEKS * 7

    # Row 1: Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    t = ws.cell(row=1, column=1, value=f"Emp: {EMP}    |    Period: {week_dates[0].strftime('%d %b %Y')} – {period_end.strftime('%d %b %Y')}")
    t.font = TITLE_FONT
    t.alignment = Alignment(horizontal="left")

    # Row 2: Week headers (merged 7 cols each, after the 2 label cols)
    for wi in range(N_WEEKS):
        col_start = 3 + wi * 7
        col_end   = col_start + 6
        ws.merge_cells(start_row=2, start_column=col_start, end_row=2, end_column=col_end)
        c = ws.cell(row=2, column=col_start, value=f"Week {wi + 1}")
        c.font = WEEK_HEADER_FONT
        c.fill = WEEK_HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Row 3: Sub-headers
    ws.cell(row=3, column=1, value="Service").font = SUB_HEADER_FONT
    ws.cell(row=3, column=2, value="Shift Type").font = SUB_HEADER_FONT
    ws.cell(row=3, column=1).fill = SUB_HEADER_FILL
    ws.cell(row=3, column=2).fill = SUB_HEADER_FILL
    ws.cell(row=3, column=1).border = CELL_BORDER
    ws.cell(row=3, column=2).border = CELL_BORDER
    for di, d in enumerate(week_dates):
        col = 3 + di
        c = ws.cell(row=3, column=col, value=d.strftime("%d-%b"))
        c.font = SUB_HEADER_FONT
        c.fill = SUB_HEADER_FILL
        c.border = CELL_BORDER
        c.alignment = Alignment(horizontal="center")

    # Body rows
    body_start = 4
    for ri, key in enumerate(sorted_keys):
        location, shift_type = key
        row_num = body_start + ri
        loc_cell = ws.cell(row=row_num, column=1, value=location)
        st_cell  = ws.cell(row=row_num, column=2, value=shift_type)
        for cell in (loc_cell, st_cell):
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = CELL_BORDER
        for di, d in enumerate(week_dates):
            col = 3 + di
            hrs = pivot[key].get(d, 0)
            c = ws.cell(row=row_num, column=col, value=round(hrs, 2) if hrs else None)
            c.border = CELL_BORDER
            c.alignment = Alignment(horizontal="center")
            if hrs:
                c.number_format = "0.00"

    # Totals row: daily totals per column, week totals at end of each week is shown as col sum below
    total_row = body_start + len(sorted_keys)
    tl = ws.cell(row=total_row, column=1, value="Daily Total")
    tl.font = TOTAL_FONT
    tl.fill = TOTAL_FILL
    tl.alignment = Alignment(horizontal="right")
    tl.border = CELL_BORDER
    ws.cell(row=total_row, column=2).fill = TOTAL_FILL
    ws.cell(row=total_row, column=2).border = CELL_BORDER

    daily_totals = []
    for di, d in enumerate(week_dates):
        col = 3 + di
        total = sum(pivot[k].get(d, 0) for k in sorted_keys)
        daily_totals.append(total)
        c = ws.cell(row=total_row, column=col, value=round(total, 2) if total else None)
        c.font = TOTAL_FONT
        c.fill = TOTAL_FILL
        c.alignment = Alignment(horizontal="center")
        c.border = CELL_BORDER
        if total:
            c.number_format = "0.00"

    # Weekly totals row
    wk_total_row = total_row + 1
    wl = ws.cell(row=wk_total_row, column=1, value="Weekly Total")
    wl.font = TOTAL_FONT
    wl.fill = PatternFill("solid", fgColor="FFE699")
    wl.alignment = Alignment(horizontal="right")
    wl.border = CELL_BORDER
    ws.cell(row=wk_total_row, column=2).fill = PatternFill("solid", fgColor="FFE699")
    ws.cell(row=wk_total_row, column=2).border = CELL_BORDER

    for wi in range(N_WEEKS):
        col_start = 3 + wi * 7
        col_end   = col_start + 6
        wk_total = sum(daily_totals[wi * 7:(wi + 1) * 7])
        ws.merge_cells(start_row=wk_total_row, start_column=col_start, end_row=wk_total_row, end_column=col_end)
        c = ws.cell(row=wk_total_row, column=col_start, value=round(wk_total, 2))
        c.font = TOTAL_FONT
        c.fill = PatternFill("solid", fgColor="FFE699")
        c.alignment = Alignment(horizontal="center")
        c.border = CELL_BORDER
        c.number_format = "0.00"

    # Column widths
    ws.column_dimensions["A"].width = 36   # Service
    ws.column_dimensions["B"].width = 22   # Shift Type
    for ci in range(3, total_cols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 8

    # Freeze panes — keep first two label cols + 3 header rows visible
    ws.freeze_panes = "C4"

    out = f"{re.sub(r'[^A-Za-z0-9]+', '_', EMP).strip('_')}_weekly_pivot.xlsx"
    wb.save(out)
    grand = sum(daily_totals)
    print(f"Wrote {out}")
    print(f"Rows (Service x Shift Type combos): {len(sorted_keys)}   "
          f"Grand total: {grand:.2f}h")


if __name__ == "__main__":
    main()
